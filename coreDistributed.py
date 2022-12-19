from openpyxl import Workbook
import math
import os

file = ''.split()
filepath = ''
wb = Workbook()
wb.create_sheet('core')
count =0
for name in file:
    count +=1
    filename = filepath + name
    
    core ={}
    maxCore = 0
    with open(f'./result/{name}/{name}-coreDistributed.txt') as f:
        for line in f:
            a,b=line.split()
            if int(b) not in core:
                core[int(a)]=int(b)
            maxCore=max(maxCore,int(a))
    core_c =[]
    for i in range(maxCore+1):
        core_c.append(0)
    for key,value in core.items():
        core_c[key]=value
    for i in range(1,maxCore+1):
        core_c[i]+=core_c[i-1]
    for i in range(maxCore+1):
        core_c[i]=core_c[i]/core_c[maxCore]
    
    core_c.pop(0)
    ws=wb['core']
    for key,value in enumerate(core_c, 1):
        print(key,value)
        ws.cell(key,count).value=value

    print(name)
wb.save(f'./result/deg-car-coreV-coreE.xlsx')